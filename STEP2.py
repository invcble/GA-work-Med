import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Load the Excel file
file_path = 'Complete_Survey_Results2.xlsx'  # Update this with your file path
df = pd.read_excel(file_path)

# Identify columns related to comments (assuming they contain the word "comment" or similar)
comment_columns = [col for col in df['Section-Question'].unique() if 'comment' in col.lower()]

# Separate the comments and non-comments data
comments_df = df[df['Section-Question'].isin(comment_columns)]
non_comments_df = df[~df['Section-Question'].isin(comment_columns)]

# Ensure the following columns are included in both DataFrames
metadata_columns = ['Survey Designator', 'Received Date', 'Service Date', 'Unit', 'Specialty']

# Pivot the data to get each barcode as a row and each Section-Question as a column
pivot_comments_df = comments_df.pivot_table(
    index=['Barcode'] + metadata_columns,
    columns='Section-Question', 
    values='Answer', 
    aggfunc=lambda x: ' '.join(str(v) for v in x if pd.notnull(v))
).reset_index()

pivot_non_comments_df = non_comments_df.pivot_table(
    index=['Barcode'] + metadata_columns,
    columns='Section-Question', 
    values='Answer', 
    aggfunc=lambda x: ' '.join(str(v) for v in x if pd.notnull(v))
).reset_index()

# Flatten the multi-index columns
pivot_comments_df.columns.name = None
pivot_non_comments_df.columns.name = None

# Save the transformed data to new Excel files
comments_output_file_path = 'Processed_Survey_Comments.xlsx'  # Update this with your desired output file path
non_comments_output_file_path = 'Processed_Survey_Non_Comments.xlsx'  # Update this with your desired output file path

# Save and format comments DataFrame
pivot_comments_df.to_excel(comments_output_file_path, index=False)

# Open the comments file with openpyxl to apply formatting
comments_wb = openpyxl.load_workbook(comments_output_file_path)
comments_ws = comments_wb.active

# Set column width to 24 and wrap text for all cells
for col in range(1, comments_ws.max_column + 1):
    column_letter = get_column_letter(col)
    comments_ws.column_dimensions[column_letter].width = 24
    for cell in comments_ws[column_letter]:
        cell.alignment = Alignment(wrap_text=True)

comments_wb.save(comments_output_file_path)

# Save and format non-comments DataFrame
pivot_non_comments_df.to_excel(non_comments_output_file_path, index=False)

# Open the non-comments file with openpyxl to apply formatting
non_comments_wb = openpyxl.load_workbook(non_comments_output_file_path)
non_comments_ws = non_comments_wb.active

# Set column width to 125 and wrap text only for the first row
for col in range(1, non_comments_ws.max_column + 1):
    column_letter = get_column_letter(col)
    non_comments_ws.column_dimensions[column_letter].width = 18
    non_comments_ws[column_letter + '1'].alignment = Alignment(wrap_text=True)

non_comments_wb.save(non_comments_output_file_path)

print(f"Comments saved in '{comments_output_file_path}' with formatting.")
print(f"Non-comments saved in '{non_comments_output_file_path}' with formatting.")
