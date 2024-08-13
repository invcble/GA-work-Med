import fitz
import re
import os
import openpyxl
from collections import defaultdict

def extract_sections(text):
    sections = re.split(r'\(bold\)(.*?)\(bold\)', text)[1::2]
    section_contents = re.split(r'\(bold\)(.*?)\(bold\)', text)[2::2]
    return dict(zip(sections, section_contents))

def extract_qa_pairs(section_content):
    lines = section_content.strip().split('\n')
    qa_pairs = []
    current_question = ""
    for line in lines:
        line = line.strip()
        if re.match(r'^\d+\.', line):
            if current_question:
                qa_pairs.append((current_question, ""))
            current_question = line
        elif current_question:
            qa_pairs.append((current_question, line))
            current_question = ""
    if current_question:
        qa_pairs.append((current_question, ""))
    return qa_pairs

def process_survey(survey_text):
    sections = extract_sections(survey_text)
    survey_data = defaultdict(list)
    
    for section, content in sections.items():
        qa_pairs = extract_qa_pairs(content)
        for question, answer in qa_pairs:
            survey_data[section].append((question, answer))
    
    return survey_data

def extract_text_from_pdf(pdf_path):
    whole_text = ""
    doc = fitz.open(pdf_path)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        blocks = page.get_text("dict")["blocks"]

        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"]
                        if "bold" in span["font"].lower():
                            whole_text += "(bold)" + text + "(bold)"
                        else:
                            whole_text += text
                    whole_text += "\n"

    whole_text = whole_text.replace("© 2023 Press Ganey Associates LLC", "")
    whole_text = whole_text.replace("† Custom Question", "")
    whole_text = whole_text.replace("†", "")
    whole_text = whole_text.replace("^ Focus Question", "")
    whole_text = whole_text.replace("^", "")

    return whole_text

# Main script
current_path = os.getcwd()
pdf_files = [file for file in os.listdir(current_path) if file.endswith(".pdf")]

workbook = openpyxl.Workbook()
sheet = workbook.active

row = 2
headers = []
header_row_written = False

for pdf_file in pdf_files:
    whole_text = extract_text_from_pdf(os.path.join(current_path, pdf_file))
    survey_array = whole_text.split("Client Name:")
    survey_array.pop(0)  # Remove any text before the first survey

    for survey_text in survey_array:
        survey_data = process_survey("Client Name:" + survey_text)
        
        if not header_row_written:
            col = 1
            for section, qa_pairs in survey_data.items():
                for question, _ in qa_pairs:
                    headers.append(f"{section}: {question}")
                    sheet.cell(row=1, column=col, value=f"{section}: {question}")
                    col += 1
            header_row_written = True
        
        col = 1
        for section, qa_pairs in survey_data.items():
            for _, answer in qa_pairs:
                sheet.cell(row=row, column=col, value=answer)
                col += 1
        row += 1

workbook.save("Survey_Results_Grouped_By_Section.xlsx")
