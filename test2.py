import pdfplumber
import fitz
import re
import os
import sys
import openpyxl


if getattr(sys, 'frozen', False):
    current_path = os.path.dirname(sys.executable)
else:
    current_path = os.getcwd()

print("Current working path is", current_path)
pdf_files = []

for file in os.listdir(current_path):
    if file.endswith(".pdf"):
        pdf_files.append(file)

print(len(pdf_files), "PDFs found!")
print("Processing PDFs, Please wait!")


# knownIDs = ["IZ0101", "IZ0101U", "HZ0101UE", "HZ0101U", "HZ0101"]
# survey_count = {}

workbook = openpyxl.Workbook()
sheet = workbook.active
global_index = 0

for pdf_file in pdf_files:
    
    ##################################
    # wholeText = ""
    # with pdfplumber.open(pdf_file) as pdf:
    #     for page in pdf.pages:

    #         wholeText += page.extract_text()

    whole_text = ""
    doc = fitz.open(pdf_file)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        blocks = page.get_text("dict")["blocks"]

        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"]
                        if "bold" in span["font"].lower():
                            whole_text += "(bold)" + text + "(bold)"
                        else:
                            whole_text += text
                    whole_text += "\n"

    wholeText = whole_text.replace("© 2023 Press Ganey Associates LLC", "")
    wholeText = wholeText.replace("† Custom Question", "")
    wholeText = wholeText.replace("†", "")
    wholeText = wholeText.replace("^ Focus Question", "")
    wholeText = wholeText.replace("^", "")

    surveyArray = wholeText.split("Client Name:")
    surveyArray.pop(0)

    for filtered_text in surveyArray:

        ##################################
        global_index += 1

        section_headers = [
            ["(bold)Background Questions(bold)", 0, 0],
            ["(bold)Admission(bold)", 0, 0],
            ["(bold)Room(bold)", 0, 0],
            ["(bold)Meals(bold)", 0, 0],
            ["(bold)Nurses(bold)", 0, 0],
            ["(bold)Tests and Treatments(bold)", 0, 0],
            ["(bold)Visitors and Family(bold)", 0, 0],
            ["(bold)Doctors(bold)", 0, 0],
            ["(bold)Discharge(bold)", 0, 0],
            ["(bold)Personal Issues(bold)", 0, 0],
            ["(bold)Overall Assessment(bold)", 0, 0],
            ["(bold)About You(bold)", 0, 0],
            ["(bold)Comm w/ Nurses(bold)", 0, 0],
            ["(bold)Response of Hosp Staff(bold)", 0, 0],
            ["(bold)Comm w/ Doctors(bold)", 0, 0],
            ["(bold)Hospital Environment(bold)", 0, 0],
            ["(bold)Communication About Pain(bold)", 0, 0],
            ["(bold)Comm About Medicines(bold)", 0, 0],
            ["(bold)Discharge Information(bold)", 0, 0],
            ["(bold)Care Transitions(bold)", 0, 0],
            ["(bold)Comments(bold)", 0, 0],
            ["Patient Name:", 0, 0]
            ]


        for i in range(len(section_headers)-1):
            if filtered_text.find(section_headers[i][0]) != -1:
                section_headers[i][1] = filtered_text.find(section_headers[i][0])
                j = i
                while filtered_text.find(section_headers[j+1][0]) == -1:
                    j += 1
                    print('Skipping section')
                section_headers[i][2] = filtered_text.find(section_headers[j+1][0])


        #####
        # print('-' * 40)


        for i in range(len(section_headers)-1):
            if global_index == 1:
                sheet.cell(row=5, column=i+5, value=section_headers[i][0].replace("(bold)",""))

            print(section_headers[i])
            # sheet.cell(row=global_index + 6, column=i+5, value=f'{section_headers[i][1]} to {section_headers[i][2]}')
            
            cell_text = filtered_text[section_headers[i][1]:section_headers[i][2]]
            if i == 0:
                # sheet.cell(row=global_index + 6, column=i+5, value=cell_text[cell_text.find(chec1)+19:cell_text.find(chec2)])
                
                cell_text = cell_text.replace('Using any number from 0 to 10, where 0 is the worst hospital', '')
                cell_text = cell_text.replace('possible and 10 is the best hospital possible, what', '')
                cell_text = cell_text.replace('number would you use to', '')
                cell_text = cell_text.replace('rate this hospital?', '')
                cell_text = cell_text.replace('Would you recommend', '')
                cell_text = cell_text.replace('this hospital', '')
                cell_text = cell_text.replace('to your friends and family?', '')
                sheet.cell(row=global_index + 6, column=i+5, value=cell_text)
                # if cell_text.find(chec) != -1:

                #     sheet.cell(row=global_index + 6, column=i+5, value="boooooom")
            else:
                sheet.cell(row=global_index + 6, column=i+5, value=cell_text)

            column_letter = openpyxl.utils.get_column_letter(i+5)
            sheet.column_dimensions[column_letter].width = 60

        print(global_index)
workbook.save("Temp.xlsx")


#1. add identifiers
#2. break down particular section
# "(bold)Background Questions(bold)
# 1. Using any number from 0 to 10, where 0 is the worst hospital possible and 10 is the best hospital possible, what 
# number would you use to rate this hospital?
# 6
# 2. Would you recommend this hospital to your friends and family?
# PROBABLY YES
# "141 205 214 218
# -1 # -1 # 70
# 173 # -1 # -1

#1. find all combinations by .replace(known question, '')
#2. for each known question, string[string.find(known)+len : string.find(known)+len + 2]