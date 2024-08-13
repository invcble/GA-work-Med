import pdfplumber
import fitz
import re
import os
import sys
import openpyxl
from collections import defaultdict

def extract_metadata(text):
    metadata = {}
    lines = text.split('\n')
    for line in lines[:10]:  # Assume metadata is in the first 10 lines
        if ':' in line:
            key, value = line.split(':', 1)
            metadata[key.strip()] = value.strip()
    return metadata

def extract_sections(text):
    sections = re.split(r'\(bold\)(.*?)\(bold\)', text)[1::2]
    section_contents = re.split(r'\(bold\)(.*?)\(bold\)', text)[2::2]
    return dict(zip(sections, section_contents))

def extract_qa_pairs(section_content):
    lines = section_content.strip().split('\n')
    qa_pairs = []
    current_question = ""
    for line in lines:
        line = line.strip()
        if re.match(r'^\d+\.', line):
            if current_question:
                qa_pairs.append((current_question, ""))
            current_question = line
        elif current_question:
            qa_pairs.append((current_question, line))
            current_question = ""
    if current_question:
        qa_pairs.append((current_question, ""))
    return qa_pairs

def process_survey(survey_text):
    metadata = extract_metadata(survey_text)
    sections = extract_sections(survey_text)
    survey_data = defaultdict(list)
    
    for section, content in sections.items():
        qa_pairs = extract_qa_pairs(content)
        for question, answer in qa_pairs:
            survey_data[section].append((question, answer))
    
    return metadata, survey_data

def extract_text_from_pdf(pdf_path):
    whole_text = ""
    doc = fitz.open(pdf_path)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        blocks = page.get_text("dict")["blocks"]

        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"]
                        if "bold" in span["font"].lower():
                            whole_text += "(bold)" + text + "(bold)"
                        else:
                            whole_text += text
                    whole_text += "\n"

    whole_text = whole_text.replace("© 2023 Press Ganey Associates LLC", "")
    whole_text = whole_text.replace("† Custom Question", "")
    whole_text = whole_text.replace("†", "")
    whole_text = whole_text.replace("^ Focus Question", "")
    whole_text = whole_text.replace("^", "")

    return whole_text

# Main script
if getattr(sys, 'frozen', False):
    current_path = os.path.dirname(sys.executable)
else:
    current_path = os.getcwd()

print("Current working path is", current_path)
pdf_files = [file for file in os.listdir(current_path) if file.endswith(".pdf")]

print(len(pdf_files), "PDFs found!")
print("Processing PDFs, Please wait!")

workbook = openpyxl.Workbook()
sheet = workbook.active

# Headers
headers = ['Survey ID', 'Client Name', 'Site Name', 'Barcode', 'Mode', 'Survey Designator', 
           'Received Date', 'Service Date', 'Unit', 'Specialty', 'Section', 'Question', 'Answer']
for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)

row = 2
survey_id = 1

for pdf_file in pdf_files:
    whole_text = extract_text_from_pdf(os.path.join(current_path, pdf_file))
    survey_array = whole_text.split("Client Name:")
    survey_array.pop(0)  # Remove any text before the first survey

    for survey_text in survey_array:
        metadata, survey_data = process_survey("Client Name:" + survey_text)
        
        for section, qa_pairs in survey_data.items():
            for question, answer in qa_pairs:
                sheet.cell(row=row, column=1, value=survey_id)
                for col, key in enumerate(['Client Name', 'Site Name', 'Barcode', 'Mode', 'Survey Designator', 
                                           'Received Date', 'Service Date', 'Unit', 'Specialty'], start=2):
                    sheet.cell(row=row, column=col, value=metadata.get(key, ''))
                sheet.cell(row=row, column=11, value=section)
                sheet.cell(row=row, column=12, value=question)
                sheet.cell(row=row, column=13, value=answer)
                row += 1
        
        survey_id += 1
        print(f"Processed survey {survey_id - 1}")

# Adjust column widths
for column in sheet.columns:
    max_length = 0
    column_letter = openpyxl.utils.get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column_letter].width = adjusted_width

workbook.save("Complete_Survey_Results.xlsx")
print("Processing complete. Results saved in 'Complete_Survey_Results.xlsx'")